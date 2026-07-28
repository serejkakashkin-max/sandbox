from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font


def _autosize_worksheet(ws):
    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_length:
                max_length = len(value)
        ws.column_dimensions[column_letter].width = min(max_length + 2, 60)


def _write_sheet(ws, headers, rows):
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    _autosize_worksheet(ws)


def build_releases_excel(records):
    wb = Workbook()
    ws = wb.active
    ws.title = "releases"

    headers = [
        "release_name",
        "project",
        "source",
        "release_id",
        "base_rov_key",
        "release_status",
        "rov_status",
        "owner",
        "start_date",
        "end_date",
        "release_date",
        "month_key",
        "year",
    ]

    rows = []
    for r in records:
        rows.append([
            r.get("release_name", ""),
            r.get("project", ""),
            r.get("source", ""),
            r.get("release_id", ""),
            r.get("base_rov_key", ""),
            r.get("release_status", ""),
            r.get("rov_status", ""),
            r.get("owner", ""),
            r.get("start_date", ""),
            r.get("end_date", ""),
            r.get("release_date", ""),
            r.get("month_key", ""),
            r.get("year", ""),
        ])

    _write_sheet(ws, headers, rows)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def build_monthly_summary_excel(monthly_series):
    wb = Workbook()
    ws = wb.active
    ws.title = "monthly_summary"

    headers = ["month", "releases_count"]
    labels = monthly_series.get("labels", [])
    values = monthly_series.get("values", [])

    rows = []
    max_len = max(len(labels), len(values))
    for i in range(max_len):
        rows.append([
            labels[i] if i < len(labels) else "",
            values[i] if i < len(values) else 0,
        ])

    _write_sheet(ws, headers, rows)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def build_full_raw_snapshot_excel(records):
    wb = Workbook()
    ws = wb.active
    ws.title = "raw_snapshot"

    all_keys = set()
    for r in records:
        raw = r.get("raw") or {}
        all_keys.update(raw.keys())

    headers = sorted(all_keys)

    rows = []
    for r in records:
        raw = r.get("raw") or {}
        row = []
        for key in headers:
            value = raw.get(key, "")
            if isinstance(value, (list, dict)):
                value = str(value)
            row.append(value)
        rows.append(row)

    _write_sheet(ws, headers, rows)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output