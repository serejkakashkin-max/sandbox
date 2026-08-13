from __future__ import annotations

import json
import re
import zipfile
from dataclasses import dataclass, field
from datetime import date, datetime, time
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


SUPPORTED_EXTENSIONS = {".txt", ".xlsx"}
KNOWN_COLUMNS = {
    "message",
    "classname",
    "servereventdatetime",
    "servicename",
    "userlogin",
    "loglevel",
}


def _normalized_key(value: Any) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value).strip().casefold())


def _serialized_cell(value: Any) -> Any:
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    return value


@dataclass(frozen=True)
class LogRecord:
    source_name: str
    row_number: int
    data: dict[str, Any]
    sheet_name: str | None = None
    _index: dict[str, Any] = field(init=False, repr=False, compare=False)

    def __post_init__(self) -> None:
        object.__setattr__(
            self,
            "_index",
            {_normalized_key(key): value for key, value in self.data.items()},
        )

    def get(self, name: str, default: Any = None) -> Any:
        value = self._index.get(_normalized_key(name), default)
        return default if value is None else value

    @property
    def location(self) -> str:
        if self.sheet_name:
            return f"{self.source_name} · {self.sheet_name} · строка {self.row_number}"
        return f"{self.source_name} · строка {self.row_number}"


@dataclass(frozen=True)
class ParseIssue:
    source_name: str
    message: str
    row_number: int | None = None


@dataclass
class ParseResult:
    records: list[LogRecord] = field(default_factory=list)
    issues: list[ParseIssue] = field(default_factory=list)
    file_stats: list[dict[str, Any]] = field(default_factory=list)

    def extend(self, other: "ParseResult") -> None:
        self.records.extend(other.records)
        self.issues.extend(other.issues)
        self.file_stats.extend(other.file_stats)


class LogParseError(ValueError):
    pass


def _decode_text(payload: bytes) -> tuple[str, str]:
    for encoding in ("utf-8-sig", "utf-8", "cp1251"):
        try:
            return payload.decode(encoding), encoding
        except UnicodeDecodeError:
            continue
    raise LogParseError("Не удалось определить кодировку TXT-файла.")


def _parse_json_fragment(text: str) -> dict[str, Any] | None:
    stripped = text.strip().rstrip(",")
    if not stripped:
        return None
    candidates = [stripped]
    if not stripped.startswith("{"):
        candidates.append("{" + stripped + "}")
    for candidate in candidates:
        try:
            value = json.loads(candidate)
        except json.JSONDecodeError:
            continue
        if isinstance(value, dict):
            return value
    return None


def parse_txt_bytes(source_name: str, payload: bytes) -> ParseResult:
    text, encoding = _decode_text(payload)
    result = ParseResult()
    stripped = text.strip()

    if stripped.startswith("["):
        try:
            items = json.loads(stripped)
        except json.JSONDecodeError:
            items = None
        if isinstance(items, list):
            for row_number, item in enumerate(items, 1):
                if isinstance(item, dict):
                    result.records.append(LogRecord(source_name, row_number, item))
                else:
                    result.issues.append(
                        ParseIssue(source_name, "Элемент JSON-массива не является объектом.", row_number)
                    )
            result.file_stats.append(
                {
                    "name": source_name,
                    "format": "TXT / JSON-массив",
                    "encoding": encoding,
                    "records": len(result.records),
                    "skipped": len(result.issues),
                }
            )
            return result

    for row_number, line in enumerate(text.splitlines(), 1):
        if not line.strip():
            continue
        record = _parse_json_fragment(line)
        if record is None:
            result.issues.append(
                ParseIssue(source_name, "Строка не распознана как JSON-запись.", row_number)
            )
            continue
        result.records.append(LogRecord(source_name, row_number, record))

    result.file_stats.append(
        {
            "name": source_name,
            "format": "TXT / построчный JSON",
            "encoding": encoding,
            "records": len(result.records),
            "skipped": len(result.issues),
        }
    )
    return result


def _inspect_xlsx_archive(
    source_name: str,
    payload: bytes,
    *,
    max_members: int,
    max_uncompressed_bytes: int,
) -> None:
    try:
        with zipfile.ZipFile(BytesIO(payload)) as archive:
            members = archive.infolist()
            if len(members) > max_members:
                raise LogParseError(
                    f"{source_name}: слишком много объектов внутри XLSX ({len(members)})."
                )
            unpacked = sum(member.file_size for member in members)
            if unpacked > max_uncompressed_bytes:
                raise LogParseError(
                    f"{source_name}: распакованный XLSX превышает допустимый размер."
                )
            if any(member.filename.startswith(("/", "\\")) or ".." in Path(member.filename).parts for member in members):
                raise LogParseError(f"{source_name}: XLSX содержит небезопасный путь.")
    except zipfile.BadZipFile as exc:
        raise LogParseError(f"{source_name}: файл не является корректным XLSX.") from exc


def _detect_header(rows: list[tuple[Any, ...]]) -> int | None:
    best_index = None
    best_score = 0
    for index, row in enumerate(rows[:20]):
        keys = {_normalized_key(value) for value in row if value is not None}
        score = len(keys & KNOWN_COLUMNS)
        if score > best_score:
            best_index = index
            best_score = score
    return best_index if best_score else None


def parse_xlsx_bytes(
    source_name: str,
    payload: bytes,
    *,
    max_members: int = 2_000,
    max_uncompressed_bytes: int = 150 * 1024 * 1024,
    max_rows: int = 200_000,
    max_columns: int = 250,
) -> ParseResult:
    _inspect_xlsx_archive(
        source_name,
        payload,
        max_members=max_members,
        max_uncompressed_bytes=max_uncompressed_bytes,
    )
    try:
        workbook = load_workbook(BytesIO(payload), read_only=True, data_only=True)
    except Exception as exc:
        raise LogParseError(f"{source_name}: не удалось прочитать книгу Excel.") from exc

    result = ParseResult()
    total_before = 0
    try:
        for sheet in workbook.worksheets:
            issues_before = len(result.issues)
            if sheet.max_column and sheet.max_column > max_columns:
                raise LogParseError(
                    f"{source_name}: на листе «{sheet.title}» слишком много столбцов."
                )
            rows = []
            for row_index, row in enumerate(sheet.iter_rows(values_only=True), 1):
                if row_index > max_rows:
                    raise LogParseError(
                        f"{source_name}: на листе «{sheet.title}» превышен лимит строк."
                    )
                rows.append(tuple(_serialized_cell(value) for value in row))

            header_index = _detect_header(rows)
            if header_index is not None:
                header = [str(value).strip() if value is not None else "" for value in rows[header_index]]
                for row_offset, row in enumerate(rows[header_index + 1 :], header_index + 2):
                    data = {
                        key: value
                        for key, value in zip(header, row)
                        if key and value not in (None, "")
                    }
                    if data:
                        result.records.append(
                            LogRecord(source_name, row_offset, data, sheet.title)
                        )
            else:
                for row_number, row in enumerate(rows, 1):
                    non_empty = [value for value in row if value not in (None, "")]
                    if len(non_empty) != 1 or not isinstance(non_empty[0], str):
                        if non_empty:
                            result.issues.append(
                                ParseIssue(
                                    source_name,
                                    f"На листе «{sheet.title}» не найдена строка заголовков логов.",
                                    row_number,
                                )
                            )
                        continue
                    data = _parse_json_fragment(non_empty[0])
                    if data is None:
                        result.issues.append(
                            ParseIssue(
                                source_name,
                                f"Ячейка листа «{sheet.title}» не распознана как JSON-запись.",
                                row_number,
                            )
                        )
                        continue
                    result.records.append(
                        LogRecord(source_name, row_number, data, sheet.title)
                    )
            added = len(result.records) - total_before
            total_before = len(result.records)
            result.file_stats.append(
                {
                    "name": source_name,
                    "sheet": sheet.title,
                    "format": "XLSX",
                    "records": added,
                    "skipped": len(result.issues) - issues_before,
                }
            )
    finally:
        workbook.close()
    return result


def parse_file_bytes(
    source_name: str,
    payload: bytes,
    *,
    max_file_bytes: int,
    max_xlsx_members: int,
    max_xlsx_uncompressed_bytes: int,
    max_xlsx_rows: int,
    max_xlsx_columns: int,
) -> ParseResult:
    suffix = Path(source_name).suffix.casefold()
    if suffix not in SUPPORTED_EXTENSIONS:
        raise LogParseError(f"{source_name}: поддерживаются только TXT и XLSX.")
    if not payload:
        raise LogParseError(f"{source_name}: файл пуст.")
    if len(payload) > max_file_bytes:
        raise LogParseError(f"{source_name}: превышен допустимый размер файла.")
    if suffix == ".txt":
        return parse_txt_bytes(source_name, payload)
    return parse_xlsx_bytes(
        source_name,
        payload,
        max_members=max_xlsx_members,
        max_uncompressed_bytes=max_xlsx_uncompressed_bytes,
        max_rows=max_xlsx_rows,
        max_columns=max_xlsx_columns,
    )


def parse_uploads(files: Iterable[Any], config: dict[str, Any]) -> ParseResult:
    uploads = [item for item in files if item and item.filename]
    if not uploads:
        raise LogParseError("Добавьте хотя бы один файл с логами.")
    if len(uploads) > config["MAX_FILES"]:
        raise LogParseError(f"Можно загрузить не более {config['MAX_FILES']} файлов за один раз.")

    combined = ParseResult()
    for upload in uploads:
        payload = upload.stream.read(config["MAX_FILE_BYTES"] + 1)
        parsed = parse_file_bytes(
            Path(upload.filename).name,
            payload,
            max_file_bytes=config["MAX_FILE_BYTES"],
            max_xlsx_members=config["MAX_XLSX_MEMBERS"],
            max_xlsx_uncompressed_bytes=config["MAX_XLSX_UNCOMPRESSED_BYTES"],
            max_xlsx_rows=config["MAX_XLSX_ROWS"],
            max_xlsx_columns=config["MAX_XLSX_COLUMNS"],
        )
        combined.extend(parsed)
    if not combined.records:
        raise LogParseError("В загруженных файлах не найдено ни одной записи лога.")
    return combined
